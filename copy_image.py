from shutil import copyfile
from openpyxl import load_workbook, Workbook



fname = 'Data/frozen_cap.xlsx'
wb = load_workbook(fname)
sheet = wb['Sheet1']

for row in sheet.iter_rows():
    src = row[1].value
    dst = row[2].value
    copyfile(src, dst)