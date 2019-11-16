from openpyxl import load_workbook

ord_no = 103699

def pre_load():
    # Load Excel ITEMS list
    wb_item = load_workbook('Data/ITEMS.xlsx')
    item_sheet = wb_item['Sheet1']

    # Load Order list
    wb_order = load_workbook('Data/Order.xlsx')
    order_sheet = wb_order['Sheet1']

    for row in order_sheet.iter_rows():
        if row[1].value != '':
            for row2 in item_sheet.iter_rows():
                if row[1].value == row2[0].value:
                    print('Row 1: ', row[1].value)
                    print('Row 2: ', row2[0].value)
                    row[6].value = row2[3].value
                    row[7].value = row2[5].value
                    row[8].value = row2[6].value
                    row[9].value = row2[7].value
        else:
            break

    wb_order.save('Data/Order.xlsx')



def load_template():
    # Load Order list
    wb_order = load_workbook('Data/Order.xlsx')
    order_sheet = wb_order['Sheet1']

    #Load Template file
    wb_temp = load_workbook('Data/Template.xlsx')
    temp_sheet = wb_temp['Sheet1']

    '''
    for row in temp_sheet.iter_rows():
        if row[0].value == '':
            cp_row = row
            break
    '''

    cp_row = temp_sheet['A2':'AS2']

    for row2 in order_sheet.iter_rows():
        temp_row = cp_row
        print(temp_row)
        temp_row[0] = ord_no
        temp_row[1] = row2[0].value
        temp_row[2] = row2[1].value
        temp_row[3] = row2[2].value
        temp_row[4] = row2[2].value
        temp_row[5] = row2[3].value
        temp_row[6] = row2[4].value
        temp_row[7] = row2[6].value
        temp_row[8] = row2[4].value
        temp_row[16] = row2[5].value
        temp_row[10] = row2[7].value
        temp_row[23] = row2[8].value
        temp_row[34] = row2[9].value
        print(temp_row)
        temp_sheet.append(temp_row)
        break


    '''
    for row2 in order_sheet.iter_rows():
        for row3 in temp_sheet.iter_rows():
            print("itter")
            if row3[0].value == '':
                row3 = row2
                print(row3)
                row3[0].value == ord_no
                row3[1].value == row2[0].value
                row3[2].value == row2[1].value
                row3[3].value == row2[2].value
                row3[4].value == row2[2].value
                row3[5].value == row2[3].value
                row3[6].value == row2[4].value
                row3[7].value == row2[6].value
                row3[8].value == row2[4].value
                row3[16].value == row2[5].value
                row3[10].value == row2[7].value
                row3[23].value == row2[8].value
                row3[34].value == row2[9].value
                break
                
    '''

    wb_temp.save('Data/Template.xlsx')




pre_load()
load_template()

