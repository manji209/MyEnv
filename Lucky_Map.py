from difflib import SequenceMatcher
import pyodbc
from openpyxl import load_workbook
import re
import pandas as pd
import numpy as np



# Load Vendor List
vend_wb = load_workbook('Data/Vendor_Items.xlsx')
vend_sheet = vend_wb['Sheet1']

# Load La Lucky List
la_wb = load_workbook('Data/Lucky_Map_To_Vendor_1.xlsx')
la_sheet = la_wb['Found']


for la_row in la_sheet.iter_rows():
    for vend_row in vend_sheet.iter_rows():
        if la_row[1].value == vend_row[3].value:
            print(la_row)
            la_row[2].value = vend_row[4].value
            la_row[3].value = vend_row[0].value
            la_row[4].value = vend_row[6].value
            break
        else:
            continue

la_wb.save(filename="Data/Lucky_Map_To_Vendor_1.xlsx")


