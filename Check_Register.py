from difflib import SequenceMatcher
import pyodbc
from openpyxl import load_workbook


# Load Excel File for Processing
wb = load_workbook('Data/EMAIL_sorted.xlsx')
email_sheet = wb.active

# Load Check Register Workbook
register_wb = load_workbook('Data/Check_Register_2019.xlsx')
count = 1
for row in email_sheet.iter_rows():
    print("Row 1: " + str(row[0].value))
    found = False
    if count == 1:
        count += 1
        continue
    for sheet in register_wb.worksheets:
        if found:
            break
        for row2 in sheet.iter_rows():
            print("Row 2: " + str(row2[1].value))
            if str(row[0].value) in str(row2[3].value):
                row[3].value = str(row2[1].value)
                row[4].value = row2[2].value
                row[5].value = row2[3].value
                found = True
                break


wb.save(filename="Data/EMAIL_sorted.xlsx")
