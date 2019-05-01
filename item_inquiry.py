from difflib import SequenceMatcher
import pyodbc
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
import numpy as np


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

# Load Excel File for Processing
wb = load_workbook('Data/ITEM_INQUIRY.xlsx')
sheet = wb['LA LUCKY']


# Connect to SQL Server and set cursor
conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=DINHPC\SQLEXPRESS;DATABASE=pbsdataDEMO;UID=pbssqluser;PWD=Admin11')
cur = conn.cursor()


sqlstring = """SELECT item_no, item_desc_1, item_desc_2, item_prc_1
                FROM dbo.ITMFIL00"""


df = pd.read_sql(sqlstring, conn)
'''
# Prompt for start Date and End Date.  Also prompt for sheet name to process
start_date = input("Input Start Date yyyy-mm-dd:  ")
end_date = input("Input End Date yyyy-mm-dd:  ")
sheet_name = input("Enter Sheet Name:  ")

df = pd.read_sql(sqlstring, conn, params=(start_date, end_date))

# Load Excel Sheet
sheet = wb[sheet_name]
'''
index = 0
none = 0
best = 0
ratio = 0
for row in sheet.iter_rows(min_col=2, max_col=2):
    index +=1
    if none >= 3:
        break
    for cell in row:
        if (cell.value != None):
            none = 0
            for row_df in df.itertuples(index=True):
                ratio = similar(cell.value, row_df.item_desc_1.rstrip())
                if ratio > best:
                    best = ratio
                    best_row = row_df

            print(row, 'ratio: ', ratio, best_row)
            best = 0
        else:
            none += 1








