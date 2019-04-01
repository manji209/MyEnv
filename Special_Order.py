from difflib import SequenceMatcher
import pyodbc
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
import numpy as np

# Load Excel File for Processing
wb = load_workbook('Data/LAL Special Orders.xlsx')
sheet = wb['Apr\'18']

'''
# Connect to SQL Server and set cursor
conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=DINHPC\SQLEXPRESS;DATABASE=pbsdataDEMO;UID=pbssqluser;PWD=Admin11')
cur = conn.cursor()


sqlstring = """SELECT cust_no, sls_rep, invc_dat, invc_no, ext_prc, item_no
                FROM dbo.IHSLIN00
                WHERE invc_dat BETWEEN ? AND ?"""

# Prompt for start Date and End Date.  Also prompt for sheet name to process
start_date = input("Input Start Date yyyy-mm-dd:  ")
end_date = input("Input End Date yyyy-mm-dd:  ")
sheet_name = input("Enter Sheet Name:  ")

df = pd.read_sql(sqlstring, conn, params=(start_date, end_date))

# Load Excel Sheet
sheet = wb[sheet_name]
'''

for row in sheet.iter_rows(min_col=6, max_col=6):
    print(row)





def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

