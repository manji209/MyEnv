from difflib import SequenceMatcher
import pyodbc
from openpyxl import load_workbook
import datetime
import pandas as pd
import numpy as np


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def get_markup(p1, p2):
    if p1 == float(0):
        return 0
    else:
        return (float(p1) - float(p2))/ float(p2)

def get_name(num):
    switcher = {
            5: "Chi Nong",
            7: "Michelle Nong",
            11: "George Nguyen",
            12: "Terry Nguyen",
            14: "Larry Nguyen",
            15: "Linh Ung",
            16: "Pierre Bach",
            17: "Kenny Nguyen",
            18: "Phat Tran",
            19: "Minh Bui",
            20: "Sang Tran"
        }
    return switcher.get(num, "NA#")


# Load Excel File for Processing
wb = load_workbook('Data/LAL Special Orders.xlsx')
sheet = wb['MAR\'19']


# Connect to SQL Server and set cursor
conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=DINHPC\SQLEXPRESS;DATABASE=pbsdataDEMO;UID=pbssqluser;PWD=Admin11')
cur = conn.cursor()


sqlstring = """SELECT cust_no, desc_1, sls_rep, qty, invc_dat, invc_no, prc, item_no
                FROM dbo.IHSLIN00
                WHERE invc_dat BETWEEN ? AND ?"""

start_date = '2019-03-01'
end_date = '2019-04-01'


df = pd.read_sql(sqlstring, conn, params=(start_date, end_date))
'''
# Prompt for start Date and End Date.  Also prompt for sheet name to process
start_date = input("Input Start Date yyyy-mm-dd:  ")
end_date = input("Input End Date yyyy-mm-dd:  ")
sheet_name = input("Enter Sheet Name:  ")

df = pd.read_sql(sqlstring, conn, params=(start_date, end_date))

# Load Excel Sheet
sheet = wb[sheet_name]
'''
index = 0
none = 0
best = 0
ratio = 0


for row in sheet.iter_rows():
    index +=1
    if none >= 10:
        break

    if (row[5].value != None):
        if index != 3:
            none = 0
            for row_df in df.itertuples(index=True):
                ratio = similar(row[5].value, row_df.desc_1.rstrip())
                if ratio > best:
                    best = ratio
                    best_row = row_df

            # Write to Excel Sheet
            row[9].value = best_row.cust_no.rstrip()
            row[10].value = get_name(int(best_row.sls_rep))
            row[11].value = best_row.invc_dat
            row[12].value = best_row.desc_1
            row[13].value = best_row.invc_no
            #row[13].value = best_row.ext_prc * best_row.qty
            row[14].value = best_row.prc
            row[16].value = best_row.item_no

            #Get Mark Up if Sheet does not already have formula
            '''
            if row[7].value == None:
                row[15].value = 0
            else:
                row[15].value = get_markup(best_row.prc, row[7].value)
            '''

            print(row, 'ratio: ', ratio, best_row)


            best = 0
    else:
        none += 1


wb.save(filename="Data/LAL Special Orders.xlsx")
'''
for row in sheet.iter_rows(min_col=6, max_col=8):
    index +=1
    if none >= 3:
        break
    for cell in row:
        if (cell[6].value != None):
            if index != 3:
                none = 0
                for row_df in df.itertuples(index=True):
                    ratio = similar(cell.value, row_df.desc_1.rstrip())
                    if ratio > best:
                        best = ratio
                        best_row = row_df

                print(row, 'ratio: ', ratio, best_row)
                best = 0
        else:
            none += 1

'''






