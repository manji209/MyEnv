import pyodbc
import pandas as pd
import xlrd
from openpyxl import load_workbook, Workbook




# Connect to SQL Server and set cursor
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=DINHPC,52052;DATABASE=pbsdata00;UID=pbssqluser;PWD=Admin11')
#conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=LALUCKYSERVER,65181;DATABASE=pbsdata00;UID=pbssqluser;PWD=Admin11')


cur = conn.cursor()

order_no = 107560

sqlstring = """SELECT ord_no, seq_no, item_no, desc_1, qty_ord, unit_prc, unit_cost FROM dbo.LINITM00 WHERE ord_no=?"""

df = pd.read_sql(sqlstring, conn, params={order_no})

df.columns = df.columns.str.strip()

df['Ext_Price'] = ''
df['Total'] = ''

fname = 'Import/LEE2.xlsx'
wb = load_workbook(fname)
sheet = wb['Sheet1']

sqlupdate = "UPDATE dbo.LINITM00 SET disc_amt=? WHERE item_no=? AND ord_no=?"

total_qty = 0
total_sales = 0
discount = 2

#parse thru dataframe to update the discounted price and totals
for idx1, row in df.iterrows():
    for row2 in sheet.iter_rows():
        if (row.item_no.strip() == row2[0].value) and (row2[2].value != None):
            print("Sanity Check", row.item_no)
            print("row: ", row2[2].value)
            values2 = [discount, row.item_no, order_no]
            try:

                cur.execute(sqlupdate, values2)
                df.loc[idx1, 'Ext_Price'] = (row.unit_prc * .98)
                print(row.item_no, (row.unit_prc * .98))
                break
                #total_qty += row.qty_ord
                #total_sales += row.Total
            except Exception as e:
                print("error")
                print(e)
        elif (row.item_no.strip() == row2[0].value):
            df.loc[idx1, 'Ext_Price'] = row.unit_prc
            break
        else:
            continue





conn.commit()


#Calculate total sales per item
for idx2, row in df.iterrows():
    print(row.item_no, row.Ext_Price)
    df.loc[idx2, 'Total'] = row.Ext_Price * row.qty_ord

for idx3, row in df.iterrows():
    total_qty += row.qty_ord
    total_sales += row.Total


query_order = """UPDATE [dbo].[ORDHDR00]
                SET [tot_qty] =?
                ,[tot_sls_amt] =?
                ,[tot_gross_amt] =?
                WHERE [ord_no] =?"""

values3 = [total_qty, total_sales, total_sales, order_no]
cur.execute(query_order, values3)
conn.commit()


