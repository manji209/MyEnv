from difflib import SequenceMatcher
import pyodbc
from openpyxl import load_workbook
import re
import pandas as pd
import numpy as np

# Connect to SQL Server and set cursor
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=DINHPC,52052;DATABASE=pbsdata00;UID=pbssqluser;PWD=Admin11')
#conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=DINHPC\SQLEXPRESS;DATABASE=pbsdataDEMO;UID=pbssqluser;PWD=Admin11')
cur = conn.cursor()

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def strip_num(oldstr):
    newstr = re.sub('[^0-9]', "", oldstr)
    newstr = re.sub('^0+', "", newstr)
    return newstr

def get_df(df_list):
    df_items = pd.DataFrame(columns=['item_no', 'item_desc_1', 'item_desc_2', 'item_standard_cost'])
    for row_df in df_list.itertuples(index=True):
        sqlstring = """SELECT item_no, item_desc_1, item_desc_2, item_standard_cost
                                    FROM dbo.ITMFIL00
                                    WHERE item_no = ?"""
        sub_df = pd.read_sql(sqlstring, conn, params={row_df.item_no})
        df_items = df_items.append(sub_df)

    return df_items


def evaluate(df_item_list):
    best = 0
    ratio = 0
    best_row = pd.DataFrame(columns=['item_no','item_desc_1','item_desc_2','item_standard_cost'])
    for row_item in df_item_list.itertuples(index=True):
        vendor = str(row[4].value) + str(row[5].value) + str(row[6].value)
        lucky = row_item.item_desc_1.rstrip() + row_item.item_desc_2.rstrip() + str(row_item.item_standard_cost)
        ratio = similar(vendor, lucky)


        if ratio > best:
            best = ratio
            best_row = row_item

    # Write to Excel Sheet

        if best_row:
            print(best_row)
            row[9].value = best_row.item_no.rstrip()
            row[10].value = best_row.item_desc_1
            row[11].value = best_row.item_desc_2
            row[12].value = best_row.item_standard_cost


    best = 0



# Load Excel File for Processing
wb = load_workbook('Data/DataEntryMARCH.xlsx')
#sheet = wb['19-1-25']
#sheet_names = wb.sheetnames


# Initialize variables
index = 0
none = 0
best = 0
ratio = 0
prev_inv = ""
SKIP = False

#Iterate through sheets
for sheet in wb.worksheets:
    for row in sheet.iter_rows():

        if (row[2].value != None):
            if (row[2].value == prev_inv) and SKIP:
                continue
            else:
                new_inv = strip_num(str(row[2].value))

                sqlstring = """SELECT item_no
                                            FROM dbo.ITXHIS00
                                            WHERE doc_no = ?"""
                df = pd.read_sql(sqlstring, conn, params={new_inv})

            if df.empty:
                prev_inv = new_inv
                SKIP = True
                continue
            else:
                prev_inv = new_inv
                SKIP = False
                df_item_list = get_df(df)
                evaluate(df_item_list)


        else:
            wb.save(filename="Data/DataEntryMARCH.xlsx")
            break


wb.save(filename="Data/DataEntryMARCH.xlsx")
