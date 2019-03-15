import pyodbc
import openpyxl
import pandas as pd
import numpy as np


# Connect to SQL Server and set cursor
conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=MULTIMEDIAPC\SQLEXPRESS;DATABASE=pbsdataDEMO;UID=pbssqluser;PWD=Admin11')

cur = conn.cursor()

# read data
#data = pd.read_excel('Data/import_order_entery_template.xlsx')

# Open the workbook and define the worksheet
# book = xlrd.open_workbook("Data/import_order_entery_template.xlsx")
#book = xlrd.open_workbook("Out/LineItem_Import.xlsx")
#sheet = book.sheet_by_name("Sheet1")


book = openpyxl.load_workbook("Out/LineItem_Import_1.xlsx")
sheet = book.active
ord_no = 0


query = """INSERT INTO [dbo].[LINITM00] (
            [ord_no]
           ,[seq_no]
           ,[item_no]
           ,[item_desc_1]
           ,[desc_1]
           ,[desc_2]
           ,[qty_ord]
           ,[ord_unit_of_meas]
           ,[qty_to_shp]
           ,[qty_bo]
           ,[qty_ret_inv]
           ,[scrap_svc_acctpc1]
           ,[scrap_svc_acctpc2]
           ,[scrap_svc_acctmain]
           ,[scrap_svc_acctsub]
           ,[bo_cod]
           ,[unit_prc]
           ,[conv_fac]
           ,[prc_unit_of_meas]
           ,[prc_unit_used]
           ,[disc_amt]
           ,[lin_item_acct_pc1]
           ,[lin_item_acct_pc2]
           ,[commis_pct]
           ,[commis_amt]
           ,[unit_cost]
           ,[txbl_cod_1]
           ,[txbl_cod_2]
           ,[txbl_cod_3]
           ,[txbl_cod_4]
           ,[txbl_cod_5]
           ,[sel_cod]
           ,[sell_unit]
           ,[qty_conv_fac]
           ,[track_flg]
           ,[track_qty]
           ,[weight]
           ,[negotiated_cost]
           ,[estimated_cost]
           ,[ex_txbl_amt]
           ,[tax_amt]
           ,[qty_scanned]
           ,[po_lin_po_no]
           ,[po_lin_lin_no]
          ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
           ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""

# grab existing row count in the database for validation later
cur.execute("SELECT count(*) FROM dbo.LINITM00")
before_import = cur.fetchone()

for r in range(1, sheet.max_row):
    ord_no = sheet.cell(r,1).value
    seq_no = sheet.cell(r,2).value
    item_no = sheet.cell(r,3).value
    item_desc_1 = sheet.cell(r,4).value
    desc_1 = sheet.cell(r,5).value
    desc_2 = sheet.cell(r,6).value
    qty_ord = sheet.cell(r,7).value
    ord_unit_of_meas = sheet.cell(r,8).value
    qty_to_shp = sheet.cell(r,9).value
    qty_bo = sheet.cell(r,10).value
    qty_ret_inv = sheet.cell(r,11).value
    scrap_svc_acctpc1 = sheet.cell(r, 12).value
    scrap_svc_acctpc2 = sheet.cell(r, 13).value
    scrap_svc_acctmain = sheet.cell(r, 14).value
    scrap_svc_acctsub = sheet.cell(r, 15).value
    bo_cod = sheet.cell(r, 16).value
    unit_prc = sheet.cell(r, 17).value
    conv_fac = sheet.cell(r, 18).value
    prc_unit_of_meas = sheet.cell(r, 19).value
    prc_unit_used = sheet.cell(r, 20).value
    disc_amt = sheet.cell(r, 21).value
    lin_item_acct_pc1 = sheet.cell(r, 22).value
    lin_item_acct_pc2 = sheet.cell(r, 23).value
    commis_pct = sheet.cell(r, 24).value
    commis_amt = sheet.cell(r, 25).value
    unit_cost = sheet.cell(r, 26).value
    txbl_cod_1 = sheet.cell(r, 27).value
    txbl_cod_2 = sheet.cell(r, 28).value
    txbl_cod_3 = sheet.cell(r, 29).value
    txbl_cod_4 = sheet.cell(r, 30).value
    txbl_cod_5 = sheet.cell(r, 31).value
    sel_cod = sheet.cell(r, 33).value
    sell_unit = sheet.cell(r, 33).value
    qty_conv_fac = sheet.cell(r, 34).value
    track_flg = sheet.cell(r, 35).value
    track_qty = sheet.cell(r, 36).value
    weight = sheet.cell(r, 37).value
    negotiated_cost = sheet.cell(r, 38).value
    estimated_cost = sheet.cell(r, 39).value
    ex_txbl_amt = sheet.cell(r, 40).value
    tax_amt = sheet.cell(r, 41).value
    qty_scanned = sheet.cell(r, 42).value
    po_lin_po_no = sheet.cell(r, 43).value
    po_lin_lin_no = sheet.cell(r, 44).value

    # Assign values from each row
    values = (ord_no, seq_no, item_no, item_desc_1, desc_1, desc_2, qty_ord, ord_unit_of_meas, qty_to_shp, qty_bo,
              qty_ret_inv, scrap_svc_acctpc1, scrap_svc_acctpc2, scrap_svc_acctmain, scrap_svc_acctsub, bo_cod, unit_prc,
              conv_fac, prc_unit_of_meas, prc_unit_used, disc_amt, lin_item_acct_pc1, lin_item_acct_pc2, commis_pct, commis_amt,
              unit_cost, txbl_cod_1, txbl_cod_2, txbl_cod_3, txbl_cod_4, txbl_cod_5, sel_cod, sell_unit, qty_conv_fac, track_flg,
              track_qty, weight, negotiated_cost, estimated_cost, ex_txbl_amt, tax_amt, qty_scanned, po_lin_po_no, po_lin_lin_no)

    # Execute sql Query
    cur.execute(query, values)



# Commit the transaction
conn.commit()

# If you want to check if all rows are imported
cur.execute("SELECT count(*) FROM dbo.LINITM00")
result = cur.fetchone()

print((result[0] - before_import[0]))  # should be True

# Close the database connection
conn.close()